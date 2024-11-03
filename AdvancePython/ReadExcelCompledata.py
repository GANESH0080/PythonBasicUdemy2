import openpyxl

# Created workbook object to find out the workbook
wk = openpyxl.load_workbook("D:\\RobotFramework\\PythonBasicsUdemy2\\ReadExcelData.xlsx")

# Creating object on which sheet you want to work
sh = wk["ReadDataSheet-1"]
print("Will work on sheet:" +sh.title)


rows = sh.max_row
cols = sh.max_column

print("Total rows are : ", +rows)
print("Total columns are : ", +cols)

for i in range(1, rows+1) :
    for j in range(1, cols+1) :
        ce = sh.cell(i,j)
        print(ce.value)

print("+++++++++++++++++++++++++++ One more way to get the cell values ++++++++++++++++++++++++++++++++")

for r in sh['A1':'C4']:
    for c in r:
        print(c.value)