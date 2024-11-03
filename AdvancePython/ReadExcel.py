import openpyxl

# Created workbook object to find out the workbook
wk = openpyxl.load_workbook("D:\\RobotFramework\\PythonBasicsUdemy2\\ReadExcelData.xlsx")

# Print the available sheets name which is under workbook
print("All available sheets are : ", wk.sheetnames)

# Print the 1st sheets name which is under workbook
print("1st available sheets is :" +wk.sheetnames[0])

# Print the 2nd sheets name which is under workbook
print("2nd available sheets is :" +wk.sheetnames[1])

# Print the active (Default Opened) sheets name which is under workbook
print("Active Sheet is :" +wk.active.title)

# Creating object on which sheet you want to work
sh = wk["ReadDataSheet-1"]
print(sh.title)

print(sh['A3'].value)
print(sh['B4'].value)

#Using one more way you can read cell data
ce = sh.cell(2,1)
print(ce.value)


ce = sh.cell(column=2,row=3)
print(ce.value)
